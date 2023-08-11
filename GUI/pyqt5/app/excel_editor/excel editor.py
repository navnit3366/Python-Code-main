# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'ui1.ui'
#
# Created by: PyQt5 UI code generator 5.15.4
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.

import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog

import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from copy import copy

import pickle

UKM_dict= {'FORBES': '1',
 'MENWA': '2',
 'PRAMUKA UH': '3',
 'TAE KWON DO': '4',
 'CATUR': '5',
 'SOFTBALL': '6',
 'KSR PMI UH': '7',
 'LIGA FILM': '8',
 'RENANG': '9',
 'TEATER KAMPUS ': '10',
 'TENIS MEJA': '11',
 'PENCAK SILAT': '12',
 'KEMPO': '13',
 'VOLI UH': '14',
 'BASKET': '15',
 'FOTOGRAFI': '16',
 'SEPAKBOLA': '17',
 'KOPMA': '18',
 'HOCKEY': '19',
 'KARATE-DO': '20',
 'MENEMBAK': '21',
 'PERBAKIN': '22',
 'BULUTANGKIS': '23',
 'KPI UH': '24',
 'KORPALA': '25',
 'SAR': '26',
 'PSM UH': '27',
 'TENIS LAPANGAN': '28',
 'LDK MPM': '29',
 'RADIO': '30',
 'SENI TARI': '31',
 'DBI-UH': '32',
 'PANAHAN': '33',
 'PERS': '34',
 'START-UP': '35',
 'BEM-UH': '36',
 'UNHAS MUN': '37',
 'IDENTITAS': '38',
 'FORKOM JAKETOS': '39',
 'AKADEMI MAHASISWA BERPRESTASI': '40',
 'IKAB-KIP': '41'}


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowIcon(QtGui.QIcon('pen_icon.png'))
        MainWindow.resize(753, 635)

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.frame_1 = QtWidgets.QFrame(self.centralwidget)
        self.frame_1.setGeometry(QtCore.QRect(10, 10, 731, 181))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        self.frame_1.setFont(font)
        self.frame_1.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_1.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_1.setObjectName("frame_1")
        self.label = QtWidgets.QLabel(self.frame_1)
        self.label.setGeometry(QtCore.QRect(0, 30, 731, 61))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(28)
        self.label.setFont(font)
        self.label.setLineWidth(0)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")

        self.frame_input_files = QtWidgets.QFrame(self.frame_1)
        self.frame_input_files.setGeometry(QtCore.QRect(10, 100, 711, 80))
        self.frame_input_files.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_input_files.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_input_files.setObjectName("frame_input_files")
        self.file_name_input = QtWidgets.QLineEdit(self.frame_input_files)
        self.file_name_input.setGeometry(QtCore.QRect(90, 20, 351, 31))

        try:
            checkpoint = pickle.load(open("checkpoint.pickle", "rb"))

            self.file_name_input.setText(checkpoint['prev_file_input'])
            self.before_backup= checkpoint['prev_before_backup']
        except (OSError, IOError) as e:
            self.before_backup= 10

        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.file_name_input.setFont(font)
        self.file_name_input.setReadOnly(True)
        self.file_name_input.setObjectName("file_name_input")
        self.browse_files = QtWidgets.QPushButton(self.frame_input_files)
        self.browse_files.setGeometry(QtCore.QRect(470, 20, 121, 31))
        self.browse_files.clicked.connect(self.act_browse_files)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.browse_files.setFont(font)
        self.browse_files.setObjectName("browse_files")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(10, 200, 731, 401))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        self.frame.setFont(font)
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.label_2 = QtWidgets.QLabel(self.frame)
        self.label_2.setGeometry(QtCore.QRect(30, 10, 701, 61))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(22)
        self.label_2.setFont(font)
        self.label_2.setLineWidth(0)
        self.label_2.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignVCenter)
        self.label_2.setObjectName("label_2")
        self.nama_ketua_input = QtWidgets.QLineEdit(self.frame)
        self.nama_ketua_input.setGeometry(QtCore.QRect(30, 120, 291, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.nama_ketua_input.setFont(font)
        self.nama_ketua_input.setReadOnly(False)
        self.nama_ketua_input.setObjectName("nama_ketua_input")
        self.label_3 = QtWidgets.QLabel(self.frame)
        self.label_3.setGeometry(QtCore.QRect(30, 90, 291, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.frame)
        self.label_4.setGeometry(QtCore.QRect(30, 180, 291, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.frame)
        self.label_5.setGeometry(QtCore.QRect(30, 270, 291, 16))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.bantua_dana_input = QtWidgets.QLineEdit(self.frame)
        self.bantua_dana_input.setGeometry(QtCore.QRect(30, 300, 291, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.bantua_dana_input.setFont(font)
        self.bantua_dana_input.setObjectName("bantua_dana_input")
        self.ukm_combobox = QtWidgets.QComboBox(self.frame)
        self.ukm_combobox.setGeometry(QtCore.QRect(30, 210, 291, 31))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.ukm_combobox.setFont(font)
        self.ukm_combobox.setObjectName("ukm_combobox")

        # for i in list(sorted(UKM_dict.keys())):
        self.ukm_combobox.addItems(list(sorted(UKM_dict.keys())))

        self.label_6 = QtWidgets.QLabel(self.frame)
        self.label_6.setGeometry(QtCore.QRect(360, 90, 291, 21))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.kegiatan_input = QtWidgets.QTextEdit(self.frame)
        self.kegiatan_input.setGeometry(QtCore.QRect(360, 120, 361, 121))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(10)
        self.kegiatan_input.setFont(font)
        self.kegiatan_input.setObjectName("kegiatan_input")
        self.submit_data = QtWidgets.QPushButton(self.frame)
        self.submit_data.setGeometry(QtCore.QRect(420, 290, 221, 51))
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.submit_data.setFont(font)
        self.submit_data.setObjectName("submit_data")
        self.submit_data.clicked.connect(self.insert_to_excel)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 753, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Excel Editor"))
        self.label.setText(_translate("MainWindow", "Excel Editor"))
        self.file_name_input.setPlaceholderText(_translate("MainWindow", "Direktori File Excel"))
        self.browse_files.setText(_translate("MainWindow", "Cari File"))
        self.label_2.setText(_translate("MainWindow", "Edit File Excel"))
        self.nama_ketua_input.setPlaceholderText(_translate("MainWindow", "Masukkan Nama Ketua"))
        self.label_3.setText(_translate("MainWindow", "Nama Ketua"))
        self.label_4.setText(_translate("MainWindow", "UKM / NON UKM"))
        self.label_5.setText(_translate("MainWindow", "Bantuan Dana"))
        self.bantua_dana_input.setPlaceholderText(_translate("MainWindow", "Masukkan Besaran Bantuan Dana"))
        self.label_6.setText(_translate("MainWindow", "Kegiatan"))
        self.kegiatan_input.setPlaceholderText(_translate("MainWindow", "Masukkan Kegiatan"))
        self.submit_data.setText(_translate("MainWindow", "Tambahkan Ke Excel"))

    def insert_to_excel(self):
        try:
            workbook = load_workbook(self.file_name_input.text())

            try:
                bantuan_dana_= int(self.bantua_dana_input.text())
            except:
                bantuan_dana_= self.bantua_dana_input.text()

            insert_data = [self.nama_ketua_input.text(), self.ukm_combobox.currentText(), self.kegiatan_input.toPlainText(), bantuan_dana_]
            insert_sheet = UKM_dict[insert_data[1]]

            self.bantua_dana_input.clear()
            self.kegiatan_input.clear()

            idx = 1
            curr_sheet = workbook[insert_sheet]
            first_insert = False

            if curr_sheet['A2'].value == 0 and curr_sheet['D2'].value is None:
                first_insert = True

                rows_count = 0
                last_rows = 4

            else:
                rows_count = -1
                last_rows = -1

                while True:
                    if 'MAX' in str(curr_sheet[idx][0].value):
                        last_rows = idx

                        idx -= 1
                        while True:
                            if curr_sheet[idx][0].value is not None:
                                rows_count = int(curr_sheet[idx][0].value)

                                break

                            idx -= 1

                        break

                    idx += 1

                assert rows_count != -1
                assert last_rows != -1

            new_rows = rows_count + 2

            new_last_rows_num_formulas = f'=MAX(A2:A{new_rows})'
            new_last_rows_aid_formulas = f'=SUM(E2:E{new_rows})'
            insert_data = [new_rows - 1] + insert_data

            prev_cell = [curr_sheet.cell(row=new_rows - 1, column=i + 1) for i in range(5)]

            for i in range(5):
                new_cell = curr_sheet.cell(row=new_rows, column=i + 1)
                new_cell.value = insert_data[i]

                if not first_insert:
                    new_cell.font = copy(prev_cell[i].font)
                    new_cell.border = copy(prev_cell[i].border)
                    new_cell.fill = copy(prev_cell[i].fill)
                    new_cell.number_format = copy(prev_cell[i].number_format)
                    new_cell.protection = copy(prev_cell[i].protection)
                    new_cell.alignment = copy(prev_cell[i].alignment)

            if not first_insert:
                curr_sheet.cell(row=last_rows, column=1, value=new_last_rows_num_formulas)
                curr_sheet.cell(row=last_rows, column=5, value=new_last_rows_aid_formulas)

            if not first_insert:
                curr_sheet.unmerge_cells(f'B{last_rows}:D{last_rows}')
                curr_sheet.insert_rows(last_rows)

            thin = Side(border_style="thin", color="000000")
            for i in curr_sheet[last_rows] + curr_sheet[last_rows - 1]:
                i.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if not first_insert:
                curr_sheet.merge_cells(f'B{last_rows + 1}:D{last_rows + 1}')

            master_sheet = workbook['MASTER']
            search_ukm = insert_data[2]

            # Pake equasi? tidak flexibel

            for search_col in range(2, master_sheet.max_column, 4):
                for search_row in range(4, master_sheet.max_row, 3):
                    broke_loop = False

                    if search_ukm == (master_sheet.cell(row=search_row, column=search_col).value):
                        broke_loop = True
                        break

                if broke_loop:
                    break

            new_col_2_value = master_sheet.cell(row=search_row + 1, column=search_col).value.split('A')[0] + 'A' + str(
                last_rows + 1)
            master_sheet.cell(row=search_row + 1, column=search_col, value=new_col_2_value)

            new_col_3_value = master_sheet.cell(row=search_row + 1, column=search_col + 1).value.split('E')[0] + 'E' + str(
                last_rows + 1)
            master_sheet.cell(row=search_row + 1, column=search_col + 1, value=new_col_3_value)

            workbook.save(self.file_name_input.text())

            self.before_backup-= 1

            if self.before_backup<=0:
                workbook.save('backup.xlsx')
                self.before_backup= 10

            checkpoint= {'prev_file_input':self.file_name_input.text(),'prev_before_backup':self.before_backup}
            pickle.dump(checkpoint, open("checkpoint.pickle", "wb"))
        except:
            print('error')



class Main(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        self.setupUi(self)

    def act_browse_files(self):
        fname= QFileDialog.getOpenFileName(self, 'Open file', filter='Excel Files (*.xlsx)')
        self.file_name_input.setText(fname[0])

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())
